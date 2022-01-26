from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from openpyxl_custom_use.format import number, hex_color


class ColumnFormat:
    def __init__(self, worksheet, col_idx, **options):
        # working spot info
        self.current_col = worksheet.column_dimensions[get_column_letter(col_idx)]

        self.col_width = options.get("col_width", 8.38)  # 8.38
        self.col_number_format = options.get("col_number_format", number.general)
        self.col_background_color = options.get("col_background_color", hex_color.white)

        # execute
        self.__col_width__()
        self.__col_number_format__()
        self.__col_background__()

    def __col_width__(self):
        current_col = self.current_col
        col_width = self.col_width

        current_col.width = col_width

    def __col_number_format__(self):
        # deals with number format of whole column
        current_col = self.current_col
        col_number_format = self.col_number_format

        current_col.number_format = col_number_format

    def __col_background__(self):
        current_col = self.current_col
        col_background_color = self.col_background_color

        current_col.fill = PatternFill(start_color=col_background_color, fill_type="solid")
