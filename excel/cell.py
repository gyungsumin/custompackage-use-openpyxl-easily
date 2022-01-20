from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side

from openpyxl.utils import get_column_letter

from excel.format import number, hex_color


class SetStyle:
    """Set cell styles at once, work with openpyxl worksheet object

    Args:
      worksheet: openpyxl worksheet object, e.g. Workbook().create_sheet(sheet_name)
      cell: string, a cell location which is a combination of one or more alphabets and an integer, e.g. "A1"
      **options: elements to be set in the cell. elements are...
        font_size: integer-like string, e.g. 11
        font_bold: boolean
        font_color: string, hex color code
        background_color: string, hex color code
        number_format: string, excel bumber format
        alignment_hor, alignment_ver: string, e.g. fill, left, distributed, general, center, right...
        border_top_style, border_bottom_style, border_left_style, border_right_style: string, e.g. thick, thin
        border_top_color, border_bottom_color, border_left_color, border_right_color: string, hex color code

    Returns:
      None.
  """
    def __init__(self, worksheet, cell, **options):
        # working spot info
        self.cell = worksheet[cell]

        # font info
        self.font_size = str(options.get("font_size", "11"))
        self.font_bold = options.get("font_bold", False)
        self.font_color = options.get("font_color", hex_color.black)

        # background info
        self.background_color = options.get("background_color", None)

        # number format info
        self.number_format = options.get("number_format", number.general)

        # cell alignment info
        self.alignment_hor = options.get("alignment_hor", "center")
        self.alignment_ver = options.get("alignment_ver", "center")

        # border info
        self.border_top_style = options.get("border_top_style", None)
        self.border_bottom_style = options.get("border_bottom_style", None)
        self.border_left_style = options.get("border_left_style", None)
        self.border_right_style = options.get("border_right_style", None)

        self.border_top_color = options.get("border_top_color", hex_color.black)
        self.border_bottom_color = options.get("border_bottom_color", hex_color.black)
        self.border_left_color = options.get("border_left_color", hex_color.black)
        self.border_right_color = options.get("border_right_color", hex_color.black)

        # execute
        self.__font__()
        self.__background__()
        self.__number_format__()
        self.__alignment__()
        self.__border__()

    def __font__(self):
        cell = self.cell
        font_size = self.font_size
        font_bold = self.font_bold
        font_color = self.font_color

        cell.font = Font(size=font_size, bold=font_bold, color=font_color)

    def __background__(self):
        cell = self.cell
        background_color = self.background_color

        cell.fill = PatternFill(start_color=background_color, fill_type="solid")

    def __number_format__(self):
        cell = self.cell
        number_format = self.number_format

        cell.number_format = number_format

    def __alignment__(self):
        cell = self.cell
        alignment_hor = self.alignment_hor
        alignment_ver = self.alignment_ver

        cell.alignment = Alignment(horizontal=alignment_hor, vertical=alignment_ver, wrap_text=True)

    def __border__(self):
        cell = self.cell
        border_top_style = self.border_top_style
        border_bottom_style = self.border_bottom_style
        border_left_style = self.border_left_style
        border_right_style = self.border_right_style
        border_top_color = self.border_top_color
        border_bottom_color = self.border_bottom_color
        border_left_color = self.border_left_color
        border_right_color = self.border_right_color

        cell.border = Border(top=Side(border_style=border_top_style, color=border_top_color),
                             bottom=Side(border_style=border_bottom_style, color=border_bottom_color),
                             left=Side(border_style=border_left_style, color=border_left_color),
                             right=Side(border_style=border_right_style, color=border_right_color))
